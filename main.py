import asyncio
import logging
import re
import os
import time
import pandas as pd
from datetime import datetime, timedelta, timezone
from aiogram import Bot, Dispatcher, F, types
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.types import FSInputFile
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from apscheduler.schedulers.asyncio import AsyncIOScheduler

from config import (
    ADMIN_ID, BOT_TOKEN, LOG_CHANNEL_ID, TARIFFS, VIEWER_ID,
    MODERATOR_IDS, SUPER_ADMIN_ID, TAX_OFFICER_IDS, TAX_RATE,
    GOOGLE_SHEETS_SPREADSHEET_ID, GOOGLE_SERVICE_ACCOUNT_FILE, GOOGLE_SHEETS_TENTS_WORKSHEET,
    MINIAPP_URL
)
import sheets_sync
import firestore_sync
import products_sync

logging.basicConfig(level=logging.WARNING)
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())

# Часовой пояс МСК (UTC+3)
MSK_TZ = timezone(timedelta(hours=3))
REPORT_START_DATE = datetime(2026, 9, 1, tzinfo=MSK_TZ)


# === СИСТЕМА ЛОГИРОВАНИЯ ОШИБОК В TELEGRAM-КАНАЛ ===
# Стандартный logging.Handler: любой logging.error(...)/logging.exception(...) в любом
# месте кода теперь автоматически долетает в LOG_CHANNEL_ID, без ручных await send_log(...).
# Работает асинхронно через очередь, чтобы не блокировать event loop синхронным вызовом.
class TelegramErrorHandler(logging.Handler):
    def __init__(self, level=logging.ERROR):
        super().__init__(level)
        self._queue: "asyncio.Queue[str] | None" = None
        self._worker_started = False

    def _ensure_worker(self):
        if self._worker_started:
            return
        try:
            loop = asyncio.get_running_loop()
        except RuntimeError:
            return
        self._queue = asyncio.Queue()
        loop.create_task(self._worker())
        self._worker_started = True

    async def _worker(self):
        while True:
            text = await self._queue.get()
            try:
                await bot.send_message(chat_id=LOG_CHANNEL_ID, text=text, parse_mode="HTML")
            except Exception:
                pass  # Не даём ошибке логирования породить бесконечный цикл логирования ошибок

    def emit(self, record: logging.LogRecord):
        if not LOG_CHANNEL_ID:
            return
        try:
            self._ensure_worker()
            if self._queue is None:
                return
            msg = self.format(record)
            text = f"🐞 <b>[ERROR | {record.levelname}]</b>\n<code>{msg[:3500]}</code>"
            self._queue.put_nowait(text)
        except Exception:
            pass


_tg_handler = TelegramErrorHandler(level=logging.ERROR)
_tg_handler.setFormatter(logging.Formatter("%(asctime)s | %(name)s | %(message)s", datefmt="%d.%m.%Y %H:%M:%S"))
logging.getLogger().addHandler(_tg_handler)

# === РОЛЕВАЯ СИСТЕМА И ПРОВЕРКА ПРАВ ===
def is_super_admin(user_id: int) -> bool:
    return user_id == SUPER_ADMIN_ID or user_id == ADMIN_ID

def can_manage_tents(user_id: int) -> bool:
    return is_super_admin(user_id) or user_id in MODERATOR_IDS

def can_export_reports(user_id: int) -> bool:
    return is_super_admin(user_id) or user_id in TAX_OFFICER_IDS


# === ПОСТОЯННАЯ КЛАВИАТУРА ВНИЗУ ЭКРАНА (вместо системной клавиатуры ввода) ===
def generate_receipt_text(receipt_id: str, nick: str, tent_id: int, days: int, price: int, end_date_str: str) -> str:
    return (
        f"🧾 <b>ОФИЦИАЛЬНЫЙ ЧЕК ОБ ОПЛАТЕ АРЕНДЫ #{receipt_id}</b>\n"
        f"──────────────────────────\n"
        f"👤 <b>Арендатор:</b> {nick}\n"
        f"⛺ <b>Объект:</b> Палатка #{tent_id}\n"
        f"⏳ <b>Оплаченный срок:</b> {days} дн.\n"
        f"🛢️ <b>Сумма оплаты:</b> {price} нефти\n"
        f"📅 <b>Действует ДО:</b> {end_date_str} 23:59:59 (МСК)\n"
        f"──────────────────────────\n"
        f"✅ <b>СТАТУС:</b> ПРОВЕДЕНО И ПОДТВЕРЖДЕНО\n"
        f"🏛️ <i>Налоговая служба / Администрация</i>"
    )


# === ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ ЛОГИРОВАНИЯ В КАНАЛ ===
async def send_log(text: str):
    """Отправка системного события в закрытый LOG-канал"""
    if LOG_CHANNEL_ID:
        try:
            now = datetime.now(MSK_TZ).strftime("%d.%m.%Y %H:%M")
            log_message = f"🪵 <b>[LOG | {now}]</b>\n{text}"
            await bot.send_message(
                chat_id=LOG_CHANNEL_ID,
                text=log_message,
                parse_mode="HTML"
            )
        except Exception as e:
            logging.error(f"❌ Ошибка отправки в LOG-канал ({LOG_CHANNEL_ID}): {e}")


# === УВЕДОМЛЕНИЯ АДМИНИСТРАЦИИ ===
# ВАЖНО: раньше уведомления слались только на ADMIN_ID, который в config.py был
# незаполненным плейсхолдером ("1" — несуществующий Telegram ID), поэтому реальный
# админ (SUPER_ADMIN_ID) никогда их не получал. Теперь уведомления уходят сразу
# всем реальным получателям: SUPER_ADMIN_ID, ADMIN_ID (если задан отдельно) и всем MODERATOR_IDS.
def get_admin_recipients() -> list:
    ids = {SUPER_ADMIN_ID, ADMIN_ID, *MODERATOR_IDS}
    return [i for i in ids if i]


async def notify_admins(text: str, parse_mode: str = "HTML"):
    for admin_id in get_admin_recipients():
        try:
            await bot.send_message(chat_id=admin_id, text=text, parse_mode=parse_mode)
        except Exception as e:
            logging.error(f"❌ Не удалось уведомить админа {admin_id}: {e}")


async def send_request_card_to_admins(photo_id: str, caption: str, reply_markup, is_document: bool = False):
    """Отправляет карточку заявки (с чеком и кнопками Подтвердить/Отклонить) всем админам/модераторам.
    Подтверждение на одной карточке помечает заявку обработанной — остальные копии
    при попытке нажать покажут 'Эта заявка уже обработана' (это безопасно)."""
    for admin_id in get_admin_recipients():
        try:
            if is_document:
                await bot.send_document(chat_id=admin_id, document=photo_id, caption=caption, reply_markup=reply_markup)
            else:
                await bot.send_photo(chat_id=admin_id, photo=photo_id, caption=caption, reply_markup=reply_markup)
        except Exception as e:
            logging.error(f"❌ Не удалось отправить карточку заявки админу {admin_id}: {e}")


def extract_proof_file(message: types.Message):
    """Достаёт file_id доказательства оплаты из сообщения, независимо от того, прислано
    оно как сжатое 'фото' или как 'файл/документ'-изображение. Возвращает (file_id, is_document)
    либо (None, None), если в сообщении нет подходящего изображения."""
    if message.photo:
        return message.photo[-1].file_id, False
    if message.document and (message.document.mime_type or "").startswith("image/"):
        return message.document.file_id, True
    return None, None


async def send_proof(chat_id: int, photo_id: str, is_document: bool = False, caption: str = None):
    """Пересылает сохранённое доказательство оплаты корректным методом в зависимости от того,
    было оно фото или документом (иначе Telegram API вернёт ошибку неверного file_id)."""
    if is_document:
        await bot.send_document(chat_id=chat_id, document=photo_id, caption=caption)
    else:
        await bot.send_photo(chat_id=chat_id, photo=photo_id, caption=caption)


async def safe_send(user_id: int, text: str, parse_mode: str = "HTML", reply_markup=None) -> bool:
    """Единая точка отправки личных сообщений игрокам напрямую по user_id.
    Возвращает True/False вместо того, чтобы падать — раньше в разных местах кода
    отправка была разбросана по отдельным try/except с разным поведением, из-за чего
    часть сбоев (заблокировал бота, удалил аккаунт, неверный ID) либо не логировалась,
    либо вообще не обрабатывалась. Здесь же можно централизованно расширить проверку —
    например, сверяться со списком известных tg_id в таблице users."""
    if not user_id:
        logging.error("safe_send: вызван без user_id")
        return False
    try:
        await bot.send_message(chat_id=user_id, text=text, parse_mode=parse_mode, reply_markup=reply_markup)
        return True
    except Exception as e:
        logging.error(f"❌ Не удалось отправить сообщение игроку {user_id}: {e}")
        return False


# === БАЗА ДАННЫХ ===
async def is_nickname_taken(nickname: str, exclude_tg_id: int = None) -> bool:
    users = await firestore_sync.list_bot_documents("bot_users")
    for data in users:
        if str(data.get("nickname", "")).lower() == nickname.lower() and int(data.get("tg_id", 0)) != int(exclude_tg_id or 0):
            return True
    return False


async def is_blacklisted(tg_id: int) -> bool:
    doc = await firestore_sync.get_bot_document("bot_blacklist", str(tg_id))
    return bool(doc)


async def ban_user(tg_id: int, reason: str = "Нарушение правил"):
    await firestore_sync.upsert_bot_document("bot_blacklist", str(tg_id), {"tg_id": tg_id, "reason": reason})


async def unban_user(tg_id: int):
    await firestore_sync.delete_bot_document("bot_blacklist", str(tg_id))


async def save_user(tg_id, username, nickname):
    await firestore_sync.upsert_bot_document("bot_users", str(tg_id), {"tg_id": tg_id, "username": username, "nickname": nickname})


async def delete_user_db(tg_id):
    await firestore_sync.delete_bot_document("bot_users", str(tg_id))

    owned = await firestore_sync.get_user_tent(tg_id)
    if owned:
        await firestore_sync.clear_tent(owned[0])
        await sync_tent_row(owned[0])


async def get_all_users():
    users = await firestore_sync.list_bot_documents("bot_users")
    return [(item.get("tg_id"), item.get("username"), item.get("nickname")) for item in users]


async def get_user_record(tg_id: int) -> dict:
    return await firestore_sync.get_bot_document("bot_users", str(tg_id))



async def get_user_tent(tg_id):
    """Теперь читает из Firestore (единый источник данных с веб-приложением и
    Google-таблицей) вместо локальной SQLite."""
    return await firestore_sync.get_user_tent(tg_id)


async def get_tent(tent_id):
    return await firestore_sync.get_tent(tent_id)


async def get_all_tents_list():
    """Список всех 20 палаток (id, tg_id, nickname, end_date) — замена массовых
    'SELECT ... FROM tents' запросов, которые раньше шли напрямую в SQLite."""
    return await firestore_sync.get_all_tents()


async def _record_payment(tent_id, nickname, price, days, photo_id, is_document, operation_type="renew"):
    """Канонический реестр платежей — коллекция Firestore 'bot_payments'
    (когда-то писала в SQLite, отсюда историческое имя функции при переносе)."""
    today = datetime.now(MSK_TZ).strftime("%d.%m.%Y %H:%M")
    payment_id = f"{tent_id}_{photo_id}_{int(datetime.now().timestamp() * 1000)}"
    await firestore_sync.upsert_bot_document("bot_payments", payment_id, {
        "tent_id": tent_id, "nickname": nickname, "price": price, "days": days,
        "photo_id": photo_id, "pay_date": today, "is_document": int(is_document),
    })
    tent = await firestore_sync.get_tent_data(tent_id) or {}
    await firestore_sync.record_rental_history(
        tent_id, nickname, price, days, tent.get("endDate") or "",
        operation_type=operation_type,
        source="bot", tg_id=tent.get("tgId"), photo_id=photo_id,
    )


async def assign_tent_db(tent_id, tg_id, nickname, days=7, price=0, photo_id=None, is_document=False):
    end_date = await firestore_sync.assign_tent(tent_id, tg_id, nickname, days=days, price=price)
    if price:
        await _record_payment(tent_id, nickname, price, days, photo_id, is_document, "rent")
    return end_date


async def clear_tent_db(tent_id):
    await firestore_sync.clear_tent(tent_id)


async def update_tent_date_db(tent_id, new_date_str):
    await firestore_sync.update_tent_date(tent_id, new_date_str)


def tent_status_label(end_date_str) -> str:
    """Единая формулировка статуса палатки по дате окончания — используется и в
    сообщении игроку 'Моя палатка', и при синхронизации в Google Sheets."""
    if not end_date_str:
        return "⚪ Свободна"
    now_msk = datetime.now(MSK_TZ).replace(tzinfo=None)
    try:
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        days_left = (end_date - now_msk).days + 1
    except Exception:
        return "⚪ Свободна"
    if days_left < 0:
        return f"🔴 ПРОСРОЧЕНА на {abs(days_left)} дн."
    if days_left <= 2:
        return f"🟡 Заканчивается (осталось {days_left} дн.)"
    return f"🟢 Активна (осталось {days_left} дн.)"


async def sync_tent_row(tent_id: int, payments_by_tent: dict = None):
    """Подтягивает текущее состояние палатки (из Firestore) и её платежей и
    отправляет строку в Google-таблицу реестра. Вызывается после КАЖДОГО изменения
    палатки — не блокирует и не ломает основной функционал бота при сбое.

    payments_by_tent: если вызывающий код уже прочитал всю коллекцию bot_payments
    (например, при массовой синхронизации всех палаток) — передаём готовый словарь
    {tent_id: [платежи]}, чтобы НЕ вычитывать всю коллекцию заново на каждую палатку.
    Раньше это не было предусмотрено, и цикл по 20 палаткам читал всю историю
    платежей 20 раз подряд при каждом запуске — это и выжигало квоту Firestore."""
    if not sheets_sync.is_configured():
        return
    tent = await get_tent(tent_id)
    if not tent:
        return
    _, tg_id, nickname, end_date = tent

    if payments_by_tent is None:
        all_payments = await firestore_sync.list_bot_documents("rental_history")
        payments = [{"price": item.get("amount", 0), "operationAt": item.get("operationAt", 0)}
                    for item in all_payments if int(item.get("tentNum", 0)) == int(tent_id)]
    else:
        payments = payments_by_tent.get(int(tent_id), [])
    payments.sort(key=lambda item: item.get("operationAt", 0))
    last_payment = payments[-1].get("price") if payments else None
    total_paid = sum(int(item.get("price") or 0) for item in payments) or None

    await sheets_sync.sync_tent(
        tent_id=tent_id,
        status=tent_status_label(end_date),
        player=nickname,
        tg_id=tg_id,
        end_date=end_date,
        last_payment=last_payment,
        total_paid=total_paid,
    )


async def extend_tent_db(tent_id, days, price, photo_id, nickname, is_document=False):
    """Продление — бизнес-логика 'от предыдущей даты, либо от сегодня если просрочено'
    теперь живёт в firestore_sync.extend_tent (ровно та же формула, что была здесь)."""
    new_end = await firestore_sync.extend_tent(tent_id, days, price, nickname)
    if price:
        await _record_payment(tent_id, nickname, price, days, photo_id, is_document, "renew")
    return new_end


async def get_stats_offsets():
    row = await firestore_sync.get_bot_document("bot_settings", "stats_corrections") or {}
    return (row.get("oil_offset", 0), row.get("deals_offset", 0))


async def update_stats_offsets(oil_offset: int, deals_offset: int):
    await firestore_sync.upsert_bot_document("bot_settings", "stats_corrections", {"id": 1, "oil_offset": oil_offset, "deals_offset": deals_offset})


# === СОСТОЯНИЯ (FSM) ===
class BroadcastState(StatesGroup):
    waiting_for_message = State()
    waiting_for_confirmation = State()

class EditDateState(StatesGroup):
    waiting_for_new_date = State()

class EditStatsState(StatesGroup):
    waiting_for_oil = State()
    waiting_for_deals = State()

class TentClaimState(StatesGroup):
    waiting_for_tent = State()
    waiting_for_nickname = State()

class DeleteUserState(StatesGroup):
    waiting_for_reason = State()

class DMState(StatesGroup):
    waiting_for_message = State()


# === МИДДЛВАРЬ / ПРОВЕРКА НА БАН ===
@dp.message.outer_middleware()
@dp.callback_query.outer_middleware()
async def check_ban_middleware(handler, event, data):
    user = getattr(event, "from_user", None)
    if user and await is_blacklisted(user.id):
        if isinstance(event, types.CallbackQuery):
            await event.answer("🚫 Вы заблокированы в системе!", show_alert=True)
        else:
            await event.answer("🚫 Вы заблокированы в Торговой Зоне и не можете использовать бота.")
        return
    return await handler(event, data)


# === ОБРАБОТКА ОТМЕНЫ СОСТОЯНИЙ (ОБЩАЯ) ===
@dp.callback_query(F.data == "cancel_state")
async def cancel_state_handler(callback: types.CallbackQuery, state: FSMContext):
    """Все оставшиеся в боте FSM (рассылка, редактирование дат/статистики, письмо
    игроку и т.д.) — админские, поэтому отмена всегда возвращает в админ-панель."""
    await state.clear()
    await callback.answer("❌ Действие отменено.")
    if can_manage_tents(callback.from_user.id):
        await show_admin_panel(callback.message.chat.id, callback.from_user.id)


# === ЛИЧНЫЕ СООБЩЕНИЯ ОТ АДМИНИСТРАЦИИ ИГРОКУ ===
@dp.callback_query(F.data.startswith("dm_user_"))
async def dm_user_start(callback: types.CallbackQuery, state: FSMContext):
    if not can_manage_tents(callback.from_user.id):
        return

    target_id = int(callback.data.split("_")[2])

    # Проверяем, что у нас вообще есть telegram_id этого игрока в базе — если запись
    # ссылается на игрока, который никогда не запускал бота (или был удалён), писать некуда.
    row = await get_user_record(target_id)

    if not row:
        await callback.answer("❌ У этого игрока нет сохранённого Telegram ID в базе — написать ему нельзя.", show_alert=True)
        return

    nick = row.get("nickname", "Игрок")
    await state.update_data(dm_target_id=target_id, dm_target_nick=nick)
    await state.set_state(DMState.waiting_for_message)

    kb = InlineKeyboardBuilder()
    kb.button(text="❌ Отменить", callback_data="cancel_state")

    await callback.message.answer(
        f"✉️ Напишите сообщение для игрока <b>{nick}</b> следующим сообщением — оно уйдёт ему в личные "
        f"от имени бота с пометкой «от администрации».",
        reply_markup=kb.as_markup(),
        parse_mode="HTML"
    )
    await callback.answer()


@dp.message(DMState.waiting_for_message)
async def dm_user_send(message: types.Message, state: FSMContext):
    data = await state.get_data()
    target_id = data.get("dm_target_id")
    nick = data.get("dm_target_nick", "Игрок")
    await state.clear()

    if not target_id:
        await message.answer("⚠️ Сессия отправки сообщения устарела, начните заново.")
        return

    if not message.text:
        await message.answer("❌ Поддерживается только текст. Сообщение не отправлено, попробуйте снова через «✉️ Написать игроку».")
        return

    delivered = await safe_send(
        target_id,
        f"📩 <b>Сообщение от администрации:</b>\n\n{message.text}"
    )

    # Дублируем ответ в чат Mini App, чтобы игрок видел переписку прямо в приложении,
    # даже если Telegram-уведомление не доставилось (бот заблокирован и т.п.).
    try:
        await firestore_sync.add_chat_message(
            target_id,
            sender="admin",
            text=message.text,
            admin_name=message.from_user.full_name,
        )
    except Exception:
        logging.exception("❌ Не удалось записать ответ администратора в чат Mini App")

    if delivered:
        await message.answer(f"✅ Сообщение доставлено игроку {nick} (и добавлено в чат приложения).")
        await send_log(f"✉️ <b>Личное сообщение:</b> {message.from_user.full_name} → {nick} (ID: <code>{target_id}</code>)\nТекст: {message.text}")
    else:
        await message.answer(f"⚠️ Не удалось отправить игроку {nick} личным сообщением (возможно, заблокировал бота), но ответ добавлен в чат приложения — он увидит его там.")


# === /start — упрощённая точка входа: только кнопка запуска Mini App ===
@dp.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext):
    """По ТЗ бот больше не ведёт никакой пользовательской логики (регистрация,
    аренда, продление и т.д. — всё это теперь только в Mini App). /start лишь
    приветствует и даёт кнопку запуска приложения. Админ-панель по-прежнему
    доступна через отдельную команду /admin."""
    await state.clear()
    kb = InlineKeyboardBuilder()
    if MINIAPP_URL:
        kb.button(text="🚀 Открыть приложение", web_app=types.WebAppInfo(url=MINIAPP_URL))
    await message.answer(
        "👋 Добро пожаловать в Торговую Зону!\n\nАренда палаток, товары, лицензии и связь с администрацией — всё в приложении.\n\nЕсли вы уже владелец палатки, но она не привязана к Telegram, используйте /claim.",
        reply_markup=kb.as_markup(),
    )


@dp.message(Command("cancel"))
async def cmd_cancel(message: types.Message, state: FSMContext):
    """Служебная команда для сброса состояния (например, если админ передумал вводить
    новую дату или текст рассылки)."""
    cur_state = await state.get_state()
    await state.clear()
    if cur_state is None:
        await message.answer("Нечего отменять.")
        return
    await message.answer("❌ Действие отменено.")
    if can_manage_tents(message.from_user.id):
        text, markup = await render_admin_panel(message.from_user.id)
        await message.answer(text, reply_markup=markup)


@dp.callback_query(F.data == "adm_broadcast")
async def start_broadcast(callback: types.CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("❌ Рассылка доступна только Главному Админу!", show_alert=True)
        return

    await callback.message.answer(
        "📢 <b>РАССЫЛКА ОБЪЯВЛЕНИЯ</b>\n\n"
        "Отправьте текст, фото или сообщение, которое получат ВСЕ зарегистрированные игроки:",
        parse_mode="HTML"
    )
    await state.set_state(BroadcastState.waiting_for_message)


@dp.message(BroadcastState.waiting_for_message)
async def process_broadcast_msg(message: types.Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        return

    await state.update_data(msg_id=message.message_id)

    kb = InlineKeyboardBuilder()
    kb.button(text="🚀 Запустить рассылку", callback_data="confirm_broadcast")
    kb.button(text="❌ Отмена", callback_data="back_admin")
    kb.adjust(1)

    await message.answer("⚠️ <b>Подтвердите отправку объявления:</b>", reply_markup=kb.as_markup(), parse_mode="HTML")
    await state.set_state(BroadcastState.waiting_for_confirmation)


@dp.callback_query(F.data == "confirm_broadcast", BroadcastState.waiting_for_confirmation)
async def run_broadcast(callback: types.CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        return

    data = await state.get_data()
    msg_id = data.get("msg_id")

    users = await get_all_users()
    count_success = 0
    count_fail = 0

    await callback.message.edit_text("🚀 Рассылка запущена... Пожалуйста, подождите.")

    for u_id, _, _ in users:
        try:
            await bot.copy_message(chat_id=u_id, from_chat_id=callback.from_user.id, message_id=msg_id)
            count_success += 1
            await asyncio.sleep(0.05)
        except Exception:
            count_fail += 1

    await callback.message.answer(
        f"✅ <b>РАССЫЛКА ЗАВЕРШЕНА!</b>\n\n"
        f"📊 Успешно доставлено: {count_success}\n"
        f"❌ Не доставлено (заблокировали бота): {count_fail}",
        parse_mode="HTML"
    )

    await send_log(f"📢 <b>Выполнена рассылка:</b>\nУспешно: {count_success} | Ошибок: {count_fail}")
    await state.clear()


# === 📦 АВТО-БЭКАП ДАННЫХ В ЛОГ-КАНАЛ КАЖДЫЕ 3 ДНЯ ===
# ВАЖНО: раньше эта функция бэкапила локальный файл tents.db, но фактическое
# рабочее хранилище (палатки, платежи, пользователи, ЧС) давно переехало в
# Firestore — sqlite используется только один раз при первом запуске как
# источник миграции. Бэкапить нужно именно Firestore, поэтому теперь сюда
# выгружаются актуальные коллекции в единый JSON-файл. Также раньше эта
# функция вообще не была подключена к планировщику (см. main()) — джоб
# добавлен ниже.
async def send_db_backup():
    if not LOG_CHANNEL_ID:
        return
    try:
        import json

        payload = {
            "generated_at": datetime.now(MSK_TZ).strftime("%d.%m.%Y %H:%M:%S"),
            "tents": [
                {"tent_id": t_id, "tg_id": tg_id, "nickname": nick, "end_date": end}
                for t_id, tg_id, nick, end in await get_all_tents_list()
            ],
            "users": [
                {"tg_id": u_id, "username": uname, "nickname": nick}
                for u_id, uname, nick in await get_all_users()
            ],
            "payments": await firestore_sync.list_bot_documents("bot_payments"),
            "blacklist": await firestore_sync.list_bot_documents("bot_blacklist"),
        }

        backup_file_name = f"backup_{datetime.now(MSK_TZ).strftime('%Y%m%d_%H%M%S')}.json"
        with open(backup_file_name, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2, default=str)

        try:
            backup_file = FSInputFile(backup_file_name)
            await bot.send_document(
                chat_id=LOG_CHANNEL_ID,
                document=backup_file,
                caption=(
                    f"📦 <b>АВТОМАТИЧЕСКИЙ БЭКАП ДАННЫХ (Firestore)</b>\n\n"
                    f"📅 Дата: {datetime.now(MSK_TZ).strftime('%d.%m.%Y %H:%M:%S')} (МСК)\n"
                    f"⛺ Палаток: {len(payload['tents'])} | 👥 Игроков: {len(payload['users'])} | "
                    f"🧾 Платежей: {len(payload['payments'])}\n"
                    f"<i>Резервная копия формируется каждые 3 дня.</i>"
                ),
                parse_mode="HTML"
            )
        finally:
            if os.path.exists(backup_file_name):
                os.remove(backup_file_name)
    except Exception as e:
        logging.error(f"Ошибка бэкапа: {e}")


# === 🔔 АВТОМАТИЧЕСКИЕ НАПОМИНАНИЯ ИГРОКАМ В ЛС (В 12:00 МСК) ===
async def check_tent_expirations():
    tents = await get_all_tents_list()

    now_msk_date = datetime.now(MSK_TZ).date()

    for tent_id, tg_id, nick, end_date_str in tents:
        if not tg_id or not end_date_str:
            continue
        try:
            end_dt = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            days_left = (end_dt - now_msk_date).days

            if days_left == 1 and tg_id:
                kb = InlineKeyboardBuilder()
                kb.button(text="💳 Продлить сейчас", callback_data=f"renew_{tent_id}")
                
                await bot.send_message(
                    chat_id=tg_id,
                    text=(
                        f"⚠️ <b>ВНИМАНИЕ! СРОК АРЕНДЫ ИСТЕКАЕТ ЗАВТРА!</b>\n\n"
                        f"Уважаемый <b>{nick}</b>!\n"
                        f"Срок аренды <b>Палатки #{tent_id}</b> заканчивается <b>завтра ({end_dt.strftime('%d.%m.%Y')}) в 23:59:59</b>.\n\n"
                        f"Пожалуйста, свяжитесь с администрацией или продлите аренду в боте!"
                    ),
                    reply_markup=kb.as_markup(),
                    parse_mode="HTML"
                )
                await send_log(f"⏰ <b>Напоминание отправлено:</b> Игроку <code>{nick}</code> (Палатка №{tent_id})")
        except Exception as e:
            logging.error(f"Ошибка отправки уведомления для {nick}: {e}")


async def _renew_redirect(target_chat_id: int, tent_id: int):
    """Общий текст+кнопка для перенаправления игрока в Mini App на продление
    конкретной палатки. Продление всегда оформляется через Mini App (там же
    оплата/чек), бот сам аренду не продлевает."""
    kb = InlineKeyboardBuilder()
    if MINIAPP_URL:
        kb.button(text="🚀 Продлить в приложении", web_app=types.WebAppInfo(url=MINIAPP_URL))
    await bot.send_message(
        chat_id=target_chat_id,
        text=(
            f"💳 Чтобы продлить аренду <b>Палатки №{tent_id}</b>, откройте Mini App — "
            f"там доступны все тарифы и оплата."
        ),
        reply_markup=kb.as_markup(),
        parse_mode="HTML",
    )


@dp.callback_query(F.data.startswith("renew_"))
async def renew_tent_callback(callback: types.CallbackQuery):
    """Раньше кнопка 'Продлить сейчас' в напоминании об окончании аренды не имела
    обработчика вообще — нажатие ничего не делало. Теперь она ведёт в Mini App,
    где и происходит фактическое продление с оплатой."""
    tent_id = int(callback.data.split("_", 1)[1])
    await callback.answer()
    await _renew_redirect(callback.message.chat.id, tent_id)


# === ✏️ РУЧНОЕ ИЗМЕНЕНИЕ СРОКА АРЕНДЫ (АДМИН) ===
@dp.callback_query(F.data.startswith("edit_date_"))
async def start_edit_date(callback: types.CallbackQuery, state: FSMContext):
    if not can_manage_tents(callback.from_user.id):
        await callback.answer("❌ У вас нет прав для изменения сроков!", show_alert=True)
        return

    tent_id = int(callback.data.split("_")[2])
    tent = await get_tent(tent_id)

    if not tent or not tent[2]:
        await callback.answer("❌ У этой палатки нет арендатора!", show_alert=True)
        return

    await state.update_data(edit_tent_id=tent_id)

    kb = InlineKeyboardBuilder()
    kb.button(text="❌ Отменить", callback_data="cancel_state")

    await callback.message.answer(
        f"✏️ <b>Изменение даты окончания аренды для Палатки №{tent_id} ({tent[2]})</b>\n\n"
        f"Текущая дата: <code>{tent[3]}</code>\n\n"
        f"Введите новую дату в формате <b>ГГГГ-ММ-ДД</b> (например: <code>2026-08-25</code>)\n"
        f"или <b>ДД.ММ.ГГГГ</b> (например: <code>25.08.2026</code>):",
        reply_markup=kb.as_markup(),
        parse_mode="HTML"
    )
    await state.set_state(EditDateState.waiting_for_new_date)


@dp.message(EditDateState.waiting_for_new_date)
async def process_new_date(message: types.Message, state: FSMContext):
    if not can_manage_tents(message.from_user.id):
        return

    data = await state.get_data()
    tent_id = data.get("edit_tent_id")
    raw_date = message.text.strip()

    parsed_date = None
    try:
        parsed_date = datetime.strptime(raw_date, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        pass

    if not parsed_date:
        try:
            parsed_date = datetime.strptime(raw_date, "%d.%m.%Y").strftime("%Y-%m-%d")
        except ValueError:
            pass

    if not parsed_date:
        kb = InlineKeyboardBuilder()
        kb.button(text="❌ Отменить", callback_data="cancel_state")
        await message.answer(
            "❌ <b>Неверный формат даты!</b>\n"
            "Используйте формат <code>ГГГГ-ММ-ДД</code> (например <code>2026-08-25</code>) "
            "или <code>ДД.ММ.ГГГГ</code> (например <code>25.08.2026</code>). Попробуйте ещё раз:",
            reply_markup=kb.as_markup(),
            parse_mode="HTML"
        )
        return

    tent = await get_tent(tent_id)
    await update_tent_date_db(tent_id, parsed_date)
    await sync_tent_row(tent_id)

    await send_log(
        f"✏️ <b>Администратор вручную изменил дату аренды:</b>\n"
        f"Палатка №{tent_id} ({tent[2]})\n"
        f"Старая дата: <code>{tent[3]}</code>\n"
        f"Новая дата: <code>{parsed_date} 23:59:59</code>"
    )

    await message.answer(
        f"✅ Дата окончания аренды палатки №{tent_id} успешно обновлена на: <b>{parsed_date} 23:59:59</b>",
        parse_mode="HTML"
    )

    if tent[1]:
        try:
            formatted_user_date = datetime.strptime(parsed_date, "%Y-%m-%d").strftime("%d.%m.%Y")
            await bot.send_message(
                chat_id=tent[1],
                text=f"ℹ️ Администратор обновил срок аренды вашей Палатки №{tent_id}.\n📅 Новая дата окончания: {formatted_user_date} в 23:59:59"
            )
        except Exception:
            pass

    await state.clear()


# === 📊 РУЧНОЕ РЕДАКТИРОВАНИЕ СТАТИСТИКИ (ОФСЕТЫ) ===
@dp.callback_query(F.data == "edit_stats_menu")
async def edit_stats_menu(callback: types.CallbackQuery):
    if not is_super_admin(callback.from_user.id):
        await callback.answer("❌ Редактирование статистики доступно только Главным Админам!", show_alert=True)
        return

    oil_off, deals_off = await get_stats_offsets()

    kb = InlineKeyboardBuilder()
    kb.button(text="✏️ Изменить нефть (офсет)", callback_data="set_oil_offset")
    kb.button(text="✏️ Изменить продления (офсет)", callback_data="set_deals_offset")
    kb.button(text="🔙 В меню статистики", callback_data="adm_stats_menu")
    kb.adjust(1)

    text = (
        f"⚙️ <b>РУЧНАЯ КОРРЕКТИРОВКА СТАТИСТИКИ</b>\n\n"
        f"Текущая поправка нефти: <b>{oil_off} 🛢️</b>\n"
        f"Текущая поправка продлений: <b>{deals_off} шт.</b>\n\n"
        f"<i>Пример: Если хотите убавить 375 нефти, введите <code>-375</code>. "
        f"Чтобы добавить, введите число (например <code>100</code>).</i>"
    )

    await callback.message.edit_text(text, reply_markup=kb.as_markup(), parse_mode="HTML")


@dp.callback_query(F.data == "set_oil_offset")
async def set_oil_offset_start(callback: types.CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        return
    await callback.message.answer("✏️ Введите число-поправку для собранной НЕФТИ (например <code>-375</code> или <code>0</code>):", parse_mode="HTML")
    await state.set_state(EditStatsState.waiting_for_oil)


@dp.message(EditStatsState.waiting_for_oil)
async def process_oil_offset(message: types.Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        return
    try:
        val = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Введите целое число (положительное или отрицательное):")
        return

    oil_off, deals_off = await get_stats_offsets()
    await update_stats_offsets(val, deals_off)

    await send_log(f"🛠️ <b>Изменён офсет нефти:</b>\nСтарый: <code>{oil_off}</code> | Новый: <code>{val}</code>")
    await message.answer(f"✅ Поправка нефти установлена: <b>{val}</b>", parse_mode="HTML")
    await state.clear()


@dp.callback_query(F.data == "set_deals_offset")
async def set_deals_offset_start(callback: types.CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        return
    await callback.message.answer("✏️ Введите число-поправку для количества ПРОДЛЕНИЙ (например <code>-3</code> или <code>0</code>):", parse_mode="HTML")
    await state.set_state(EditStatsState.waiting_for_deals)


@dp.message(EditStatsState.waiting_for_deals)
async def process_deals_offset(message: types.Message, state: FSMContext):
    if not is_super_admin(message.from_user.id):
        return
    try:
        val = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Введите целое число (положительное или отрицательное):")
        return

    oil_off, deals_off = await get_stats_offsets()
    await update_stats_offsets(oil_off, val)

    await send_log(f"🛠️ <b>Изменён офсет продлений:</b>\nСтарый: <code>{deals_off}</code> | Новый: <code>{val}</code>")
    await message.answer(f"✅ Поправка продлений установлена: <b>{val}</b>", parse_mode="HTML")
    await state.clear()


# === 🧹 ПОЛНАЯ ОЧИСТКА БАЗЫ ДАННЫХ ===
@dp.message(Command("clear_database"))
async def clear_database_cmd(message: types.Message):
    if not is_super_admin(message.from_user.id):
        await message.answer("❌ У вас нет прав на полную очистку базы данных!")
        return

    for payment in await firestore_sync.list_bot_documents("bot_payments"):
        await firestore_sync.delete_bot_document("bot_payments", payment["id"])
    await update_stats_offsets(0, 0)

    await send_log("🧹 <b>Супер-Админ выполнил полную очистку базы платежей.</b>")
    await message.answer("🧹 <b>База данных успешно очищена!</b> Все записи, платежи и поправки удалены.")


# === 📥 ЭКСПОРТ ФИНАНСОВОГО ОТЧЁТА В EXCEL ===
@dp.callback_query(F.data == "export_excel")
async def select_excel_period(callback: types.CallbackQuery):
    if not can_export_reports(callback.from_user.id):
        await callback.answer("❌ У вас нет доступа к выгрузке отчётов Налоговой!", show_alert=True)
        return

    kb = InlineKeyboardBuilder()
    kb.button(text="📄 Отчёт за 7 дней (1 неделя)", callback_data="ex_period_7")
    kb.button(text="📄 Отчёт за 14 дней (2 недели)", callback_data="ex_period_14")
    kb.button(text="📄 Отчёт за 30 дней (1 месяц)", callback_data="ex_period_30")
    kb.button(text="📄 Отчёт за всё время", callback_data="ex_period_all")
    kb.button(text="🔙 Назад", callback_data="adm_stats_menu")
    kb.adjust(1)

    await callback.message.edit_text(
        "📊 <b>ФИНАНСОВЫЙ ОТЧЁТ ДЛЯ НАЛОГОВОЙ / ПРЕЗИДЕНТА</b>\n\n"
        "Выберите период, за который необходимо сформировать выгрузку в Excel:",
        reply_markup=kb.as_markup(),
        parse_mode="HTML"
    )


@dp.callback_query(F.data.startswith("ex_period_"))
async def process_excel_export(callback: types.CallbackQuery):
    if not can_export_reports(callback.from_user.id):
        return

    period_code = callback.data.split("_")[2]
    await callback.answer("📊 Формируем финансовую ведомость...")

    all_payments = []
    for item in await firestore_sync.list_bot_documents("rental_history"):
        operation_at = int(item.get("operationAt") or 0)
        pay_date = datetime.fromtimestamp(operation_at / 1000, MSK_TZ).strftime("%d.%m.%Y %H:%M") if operation_at else ""
        all_payments.append((item.get("tentNum"), item.get("player"), item.get("amount", 0), item.get("days", 0), pay_date, operation_at))
    all_payments.sort(key=lambda item: item[5])

    now_msk = datetime.now(MSK_TZ).replace(tzinfo=None)

    filtered_payments = []
    if period_code == "all":
        period_title = "С 01.09.2026"
        start_ms = int(REPORT_START_DATE.timestamp() * 1000)
        filtered_payments = [payment for payment in all_payments if payment[5] >= start_ms]
    else:
        days = int(period_code)
        period_title = f"За {days} дней ({days // 7} нед.)"
        limit_ms = max(
            int((now_msk - timedelta(days=days)).replace(tzinfo=MSK_TZ).timestamp() * 1000),
            int(REPORT_START_DATE.timestamp() * 1000),
        )

        for p in all_payments:
            if p[5] >= limit_ms:
                filtered_payments.append(p)

    aggregated = {}
    for tent_id, nick, price, days, pay_date, _operation_at in filtered_payments:
        key = (nick, tent_id)
        if key not in aggregated:
            aggregated[key] = {
                "nickname": nick,
                "tent_id": tent_id,
                "total_days": int(days or 0),
                "total_oil": int(price or 0),
                "last_date": pay_date,
                "deals_count": 1
            }
        else:
            aggregated[key]["total_days"] += int(days or 0)
            aggregated[key]["total_oil"] += int(price or 0)
            aggregated[key]["last_date"] = pay_date
            aggregated[key]["deals_count"] += 1

    total_oil_sum = sum(item["total_oil"] for item in aggregated.values())
    total_deals_sum = sum(item["deals_count"] for item in aggregated.values())
    total_players_count = len(aggregated)

    rows_data = []
    for idx, item in enumerate(aggregated.values(), 1):
        rows_data.append([
            idx,
            item["nickname"],
            f"Палатка #{item['tent_id']}",
            item["total_days"],
            item["total_oil"],
            item["last_date"]
        ])

    # Имя файла включает user_id и метку времени с миллисекундами — иначе при
    # параллельном экспорте двумя админами один процесс мог удалить (os.remove)
    # файл, который в этот момент ещё читает/отправляет другой (гонка).
    file_name = (
        f"Nalogovay_Otchet_{datetime.now(MSK_TZ).strftime('%d_%m_%Y')}"
        f"_{callback.from_user.id}_{int(datetime.now().timestamp() * 1000)}.xlsx"
    )

    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        columns = ["№ п/п", "Игровой Ник", "Палатка №", "Срок аренды (дней)", "Оплачено нефти (🛢️)", "Дата посл. операции"]
        df = pd.DataFrame(rows_data, columns=columns)
        df.to_excel(writer, sheet_name='Финансовый отчёт', index=False, startrow=7)

        workbook = writer.book
        worksheet = writer.sheets['Финансовый отчёт']

        font_header_title = Font(name='Arial', size=14, bold=True, color='1B365D')
        font_sub = Font(name='Arial', size=10, italic=True, color='555555')
        font_stat_label = Font(name='Arial', size=11, bold=True)
        font_stat_val = Font(name='Arial', size=11, bold=True, color='006100')
        
        font_th = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        fill_th = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')

        font_total = Font(name='Arial', size=11, bold=True)
        fill_total = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        border_total = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='double', color='000000')
        )

        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')

        worksheet['A1'] = "СВОДНАЯ ФИНАНСОВАЯ ВЕДОМОСТЬ"
        worksheet['A1'].font = font_header_title
        
        worksheet['A2'] = f"Период отчёта: {period_title} | Дата формирования: {datetime.now(MSK_TZ).strftime('%d.%m.%Y %H:%M')} (МСК)"
        worksheet['A2'].font = font_sub

        worksheet['A4'] = "ИТОГО СОБРАНО НЕФТИ (НАЛОГОВ):"
        worksheet['A4'].font = font_stat_label
        worksheet['D4'] = f"{total_oil_sum} 🛢️"
        worksheet['D4'].font = font_stat_val

        worksheet['A6'] = "ГОСУДАРСТВУ (70%) / ОСТАТОК (30%):"
        worksheet['A6'].font = font_stat_label
        worksheet['D6'] = f"{total_oil_sum * 70 // 100} 🛢️ / {total_oil_sum * 30 // 100} 🛢️"
        worksheet['D6'].font = font_stat_val

        worksheet['A5'] = "ВСЕГО АРЕНДАТОРОВ / ТРАНЗАКЦИЙ:"
        worksheet['A5'].font = font_stat_label
        worksheet['D5'] = f"{total_players_count} чел. / {total_deals_sum} сдел."
        worksheet['D5'].font = font_stat_label

        for col_num in range(1, 7):
            cell = worksheet.cell(row=8, column=col_num)
            cell.font = font_th
            cell.fill = fill_th
            cell.alignment = align_center

        data_start_row = 9
        for row_idx in range(data_start_row, data_start_row + len(rows_data)):
            worksheet.cell(row=row_idx, column=1).alignment = align_center
            worksheet.cell(row=row_idx, column=2).alignment = align_left
            worksheet.cell(row=row_idx, column=3).alignment = align_center
            worksheet.cell(row=row_idx, column=4).alignment = align_right
            worksheet.cell(row=row_idx, column=5).alignment = align_right
            worksheet.cell(row=row_idx, column=6).alignment = align_center

            for col_num in range(1, 7):
                worksheet.cell(row=row_idx, column=col_num).border = border_thin

        total_row = data_start_row + len(rows_data)
        worksheet.cell(row=total_row, column=1, value="").fill = fill_total
        worksheet.cell(row=total_row, column=2, value="ИТОГО ПО ВЕДОМОСТИ:").font = font_total
        worksheet.cell(row=total_row, column=2).fill = fill_total
        worksheet.cell(row=total_row, column=3, value="").fill = fill_total
        
        total_days_sum = sum(item["total_days"] for item in aggregated.values())
        worksheet.cell(row=total_row, column=4, value=total_days_sum).font = font_total
        worksheet.cell(row=total_row, column=4).alignment = align_right
        worksheet.cell(row=total_row, column=4).fill = fill_total
        
        worksheet.cell(row=total_row, column=5, value=total_oil_sum).font = font_total
        worksheet.cell(row=total_row, column=5).alignment = align_right
        worksheet.cell(row=total_row, column=5).fill = fill_total
        
        worksheet.cell(row=total_row, column=6, value="").fill = fill_total

        for col_num in range(1, 7):
            worksheet.cell(row=total_row, column=col_num).border = border_total

        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 15)

    try:
        excel_file = FSInputFile(file_name)
        await callback.message.answer_document(
            document=excel_file,
            caption=(
                f"🏛️ <b>СВОДНЫЙ ФИНАНСОВЫЙ ОТЧЁТ</b>\n\n"
                f"📌 <b>Период:</b> {period_title}\n"
                f"🛢️ <b>Всего собрано:</b> {total_oil_sum} 🛢️\n"
                f"👥 <b>Активных арендаторов:</b> {total_players_count}\n"
                f"🧾 <b>Всего операций:</b> {total_deals_sum}\n\n"
                f"<i>Дублирующие операции объединены. Отчёт готов для Налоговой.</i>"
            ),
            parse_mode="HTML"
        )
        await send_log(f"🏛️ <b>Сформирован сводный отчёт для Налоговой</b> за период: {period_title}")
    except Exception as e:
        await callback.message.answer(f"❌ Ошибка при отправке Excel: {e}")
    finally:
        if os.path.exists(file_name):
            os.remove(file_name)


# === АДМИН-ПАНЕЛЬ И СТАТИСТИКА ===
@dp.message(Command("stats"))
async def cmd_stats(message: types.Message):
    if can_export_reports(message.from_user.id) or message.from_user.id == VIEWER_ID:
        await show_stats_menu(message)


async def render_admin_panel(user_id: int):
    """Строит текст и inline-клавиатуру админ-панели с учётом прав конкретного user_id."""
    tents = await get_all_tents_list()

    kb = InlineKeyboardBuilder()
    for t_id, tg_id, nick, end in tents:
        btn_text = f"#{t_id} 🟢 {nick}" if nick else f"#{t_id} ⚪ Свободна"
        kb.button(text=btn_text, callback_data=f"adm_tent_{t_id}")

    kb.adjust(2)
    if is_super_admin(user_id):
        kb.row(
            types.InlineKeyboardButton(text="📢 Рассылка игрокам", callback_data="adm_broadcast"),
            types.InlineKeyboardButton(text="📊 Статистика ТЗ", callback_data="adm_stats_menu")
        )
        kb.row(
            types.InlineKeyboardButton(text="👥 Список игроков", callback_data="adm_users_list"),
            types.InlineKeyboardButton(text="🚫 Чёрный список", callback_data="adm_blacklist")
        )
    else:
        kb.row(types.InlineKeyboardButton(text="📊 Статистика ТЗ", callback_data="adm_stats_menu"))

    return "🛠️ УПРАВЛЕНИЕ ПАЛАТКАМИ И СИСТЕМОЙ:", kb.as_markup()


async def show_admin_panel(chat_id: int, user_id: int):
    """Отправляет админ-панель по chat_id, права проверяются по РЕАЛЬНОМУ user_id нажавшего
    кнопку — используйте эту функцию вместо cmd_admin(callback.message) внутри callback-хендлеров."""
    if not can_manage_tents(user_id):
        return
    text, markup = await render_admin_panel(user_id)
    await bot.send_message(chat_id=chat_id, text=text, reply_markup=markup)


@dp.message(Command("admin"))
async def cmd_admin(message: types.Message):
    if not can_manage_tents(message.from_user.id):
        return
    text, markup = await render_admin_panel(message.from_user.id)
    await message.answer(text, reply_markup=markup)


# --- ЧЁРНЫЙ СПИСОК ---
@dp.callback_query(F.data == "adm_blacklist")
async def adm_blacklist_menu(callback: types.CallbackQuery):
    if not is_super_admin(callback.from_user.id):
        return

    banned = []
    for item in await firestore_sync.list_bot_documents("bot_blacklist"):
        user = await get_user_record(int(item.get("tg_id", 0))) or {}
        banned.append((item.get("tg_id"), user.get("nickname"), user.get("username")))

    msg = "🚫 ЧЁРНЫЙ СПИСОК ИГРОКОВ:\n\n"
    kb = InlineKeyboardBuilder()

    if not banned:
        msg += "В чёрном списке пока никого нет."
    else:
        for tg_id, nick, uname in banned:
            name = nick or (f"@{uname}" if uname else f"ID: {tg_id}")
            msg += f"• {name} (ID: {tg_id})\n"
            kb.button(text=f"🟢 Разбанить {name}", callback_data=f"unban_{tg_id}")

    kb.button(text="🔙 Назад в меню", callback_data="back_admin")
    kb.adjust(1)

    await callback.message.edit_text(msg, reply_markup=kb.as_markup())


@dp.callback_query(F.data.startswith("ban_user_"))
async def process_ban_callback(callback: types.CallbackQuery):
    if not is_super_admin(callback.from_user.id):
        return

    user_id = int(callback.data.split("_")[2])
    await ban_user(user_id)

    await send_log(f"🚫 <b>Заблокирован игрок:</b> ID <code>{user_id}</code>")
    await callback.answer("🚫 Игрок заблокирован!", show_alert=True)
    await adm_users_list(callback)


@dp.callback_query(F.data.startswith("unban_"))
async def process_unban_callback(callback: types.CallbackQuery):
    if not is_super_admin(callback.from_user.id):
        return

    user_id = int(callback.data.split("_")[1])
    await unban_user(user_id)

    await send_log(f"🟢 <b>Разблокирован игрок:</b> ID <code>{user_id}</code>")
    await callback.answer("🟢 Игрок разблокирован!", show_alert=True)
    await adm_blacklist_menu(callback)


# --- СТАТИСТИКА И ПОДМЕНЮ ---
async def show_stats_menu(target):
    kb = InlineKeyboardBuilder()
    kb.button(text="📅 За 7 дней (неделя)", callback_data="st_7")
    kb.button(text="📅 За 14 дней (2 недели)", callback_data="st_14")
    kb.button(text="📅 За 30 дней (месяц)", callback_data="st_30")
    kb.button(text="♾️ За всё время", callback_data="st_all")

    user_id = target.from_user.id if hasattr(target, 'from_user') else None
    if is_super_admin(user_id):
        kb.button(text="✏️ Корректировать числа вручную", callback_data="edit_stats_menu")
    if can_export_reports(user_id):
        kb.button(text="📥 Выгрузить отчёт в Excel", callback_data="export_excel")

    if can_manage_tents(user_id):
        kb.button(text="🔙 Назад в админку", callback_data="back_admin")

    kb.adjust(1)

    text = "📊 <b>ВЫБЕРИТЕ ПЕРИОД ДЛЯ ОТЧЁТА И СТАТИСТИКИ:</b>"
    if isinstance(target, types.CallbackQuery):
        await target.message.edit_text(text, reply_markup=kb.as_markup(), parse_mode="HTML")
    else:
        await target.answer(text, reply_markup=kb.as_markup(), parse_mode="HTML")


@dp.callback_query(F.data == "adm_stats_menu")
async def adm_stats_menu_cb(callback: types.CallbackQuery):
    await show_stats_menu(callback)


@dp.callback_query(F.data.startswith("st_"))
async def process_stats_period(callback: types.CallbackQuery):
    # Раньше этот хендлер не проверял права вообще — доступ был защищён только тем,
    # что кнопка не показывалась не-админам. Но callback_data можно вызвать напрямую,
    # зная его строку, поэтому здесь нужна та же проверка, что и в show_stats_menu.
    user_id = callback.from_user.id
    if not (can_export_reports(user_id) or can_manage_tents(user_id) or user_id == VIEWER_ID):
        await callback.answer("❌ Нет доступа к статистике.", show_alert=True)
        return

    period_code = callback.data.split("_")[1]

    users_count = len(await firestore_sync.list_bot_documents("bot_users"))
    history = await firestore_sync.list_bot_documents("rental_history")
    all_payments = [
        (item.get("amount", 0), item.get("operationAt", 0))
        for item in history
    ]

    all_tents = await get_all_tents_list()
    occupied_tents = [t for t in all_tents if t[1] is not None]
    occupied_count = len(occupied_tents)

    total_earned = 0
    total_deals = 0

    now_msk = datetime.now(MSK_TZ).replace(tzinfo=None)

    if period_code == "all":
        period_title = "С 01.09.2026"
        for price, p_date in all_payments:
            if int(p_date or 0) >= int(REPORT_START_DATE.timestamp() * 1000):
                total_earned += int(price or 0)
                total_deals += 1
    else:
        days = int(period_code)
        period_title = f"За {days} дней"
        limit_date = now_msk - timedelta(days=days)

        limit_ms = max(int(limit_date.replace(tzinfo=MSK_TZ).timestamp() * 1000), int(REPORT_START_DATE.timestamp() * 1000))
        for price, operation_at in all_payments:
            if int(operation_at or 0) >= limit_ms:
                total_earned += int(price or 0)
                total_deals += 1

    oil_off, deals_off = await get_stats_offsets()
    total_earned = max(0, total_earned + oil_off)
    total_deals = max(0, total_deals + deals_off)

    expiring_list = ""
    for t_id, tg_id, nick, end_str in occupied_tents:
        try:
            end_date = datetime.strptime(end_str, "%Y-%m-%d")
            days_left = (end_date - now_msk).days + 1
            if days_left <= 2:
                expiring_list += f"\n• Палатка #{t_id} ({nick}): осталось {days_left} дн."
        except Exception:
            pass

    msg = (
        f"📊 <b>ОТЧЁТ СТАТИСТИКИ ({period_title.upper()}):</b>\n\n"
        f"🛢️ Собрано нефти: <b>{total_earned} 🛢️</b>\n"
        f"🏛️ Государству 70%: <b>{total_earned * 70 // 100} 🛢️</b>\n"
        f"🏕️ Остаётся 30%: <b>{total_earned * 30 // 100} 🛢️</b>\n"
        f"🧾 Сделок/Продлений: <b>{total_deals}</b>\n"
        f"⛺ Занято палаток: <b>{occupied_count} из 20</b> (Свободно: {20 - occupied_count})\n"
        f"👥 Всего зарегистрировано игроков: <b>{users_count}</b>\n"
    )

    if expiring_list:
        msg += f"\n⚠️ <b>Заканчивается срок аренды:</b>{expiring_list}"
    else:
        msg += "\n🟢 Срочных просрочек нет."

    kb = InlineKeyboardBuilder()
    kb.button(text="🔙 К выбору периода", callback_data="adm_stats_menu")
    if can_manage_tents(callback.from_user.id):
        kb.button(text="🔙 В главное меню", callback_data="back_admin")
    kb.adjust(1)

    await callback.message.edit_text(msg, reply_markup=kb.as_markup(), parse_mode="HTML")


@dp.callback_query(F.data.startswith("history_menu_"))
async def history_period_menu(callback: types.CallbackQuery):
    if not can_manage_tents(callback.from_user.id):
        return

    tent_id = int(callback.data.split("_")[2])

    kb = InlineKeyboardBuilder()
    kb.button(text="📅 За последний месяц (30д)", callback_data=f"show_hist_{tent_id}_month")
    kb.button(text="♾️ За всё время", callback_data=f"show_hist_{tent_id}_all")
    kb.button(text="🔙 Назад к палатке", callback_data=f"adm_tent_{tent_id}")
    kb.adjust(1)

    await callback.message.edit_text(
        f"📜 История платежей палатки №{tent_id}\nВыберите период просмотра:",
        reply_markup=kb.as_markup()
    )


@dp.callback_query(F.data.startswith("show_hist_"))
async def show_history_photos(callback: types.CallbackQuery):
    if not can_manage_tents(callback.from_user.id):
        return

    parts = callback.data.split("_")
    tent_id = int(parts[2])
    period = parts[3]

    rows = [
        (item.get("player"), item.get("amount"), item.get("days"), item.get("photoId"),
         datetime.fromtimestamp(int(item.get("operationAt") or 0) / 1000, MSK_TZ).strftime("%d.%m.%Y %H:%M"),
         0)
        for item in await firestore_sync.list_bot_documents("rental_history")
        if int(item.get("tentNum", 0)) == tent_id
    ]
    rows.sort(key=lambda row: row[4] or "", reverse=True)

    filtered_rows = []
    now_msk = datetime.now(MSK_TZ).replace(tzinfo=None)

    if period == "month":
        limit_date = now_msk - timedelta(days=30)
        for row in rows:
            pay_date_str = row[4]
            try:
                dt = datetime.strptime(pay_date_str, "%d.%m.%Y %H:%M")
                if dt >= limit_date:
                    filtered_rows.append(row)
            except Exception:
                filtered_rows.append(row)
    else:
        filtered_rows = rows

    if not filtered_rows:
        await callback.answer("❌ За выбранный период подтверждённых чеков не найдено.", show_alert=True)
        return

    await callback.answer("📤 Отправка чеков...")

    for nick, price, days, photo_id, pay_date, is_document in filtered_rows:
        caption = (
            f"🧾 Оплата Палатки №{tent_id}\n"
            f"👤 Игрок: {nick}\n"
            f"💰 Сумма: {price} 🛢️ нефти ({days} дн.)\n"
            f"📅 Дата оплаты: {pay_date}"
        )
        try:
            if photo_id:
                await send_proof(callback.from_user.id, photo_id, is_document=bool(is_document), caption=caption)
            else:
                await bot.send_message(chat_id=callback.from_user.id, text=f"{caption}\n(Фото отсутствует)")
            await asyncio.sleep(0.3)
        except Exception as e:
            logging.error(f"Ошибка вывода чека: {e}")


@dp.callback_query(F.data.startswith("adm_tent_"))
async def adm_tent_manage(callback: types.CallbackQuery):
    if not can_manage_tents(callback.from_user.id):
        return

    tent_id = int(callback.data.split("_")[2])
    tent = await get_tent(tent_id)
    _, tg_id, nickname, end_date = tent

    kb = InlineKeyboardBuilder()

    if not nickname:
        text = f"⛺ Палатка №{tent_id}\nСтатус: ⚪ Свободна"

        all_tents = await get_all_tents_list()
        occupied_tg_ids = {t[1] for t in all_tents if t[1] is not None}

        all_users = [(item[0], item[2]) for item in await get_all_users()]
        unassigned = [(u_id, u_nick) for u_id, u_nick in all_users if u_id not in occupied_tg_ids]

        if unassigned:
            text += "\n\nИгроки без палатки:"
            for u_id, u_nick in unassigned:
                kb.button(text=f"➕ Выдать {u_nick}", callback_data=f"give_{tent_id}_{u_id}")
    else:
        formatted_end = end_date or "Не установлена"
        if end_date:
            try:
                formatted_end = datetime.strptime(end_date, "%Y-%m-%d").strftime("%d.%m.%Y")
            except Exception:
                pass

        text = (
            f"⛺ Палатка №{tent_id}\n"
            f"👤 Игрок: {nickname}\n"
            f"📅 Оплачена до: {formatted_end} 23:59:59"
        )

        kb.button(text="✏️ Изменить дату аренды", callback_data=f"edit_date_{tent_id}")
        kb.button(text="📜 История платежей (Чеки)", callback_data=f"history_menu_{tent_id}")
        kb.button(text="✉️ Написать игроку", callback_data=f"dm_user_{tg_id}")
        kb.button(text="❌ Завершить аренду (Освободить)", callback_data=f"clear_{tent_id}_{tg_id}")

    kb.button(text="🔙 Назад в меню", callback_data="back_admin")
    kb.adjust(1)

    await callback.message.edit_text(text, reply_markup=kb.as_markup())


@dp.callback_query(F.data.startswith("give_"))
async def process_give(callback: types.CallbackQuery):
    if not can_manage_tents(callback.from_user.id):
        return

    _, tent_id, user_id = callback.data.split("_")
    tent_id, user_id = int(tent_id), int(user_id)

    # Ник берём СВЕЖИЙ из базы по user_id, а не из текста кнопки — раньше ник встраивался
    # прямо в callback_data и если в нём была "_" (обычное дело для Minecraft-ников,
    # например "Cool_Guy"), разбор строки падал с ошибкой и выдача палатки просто не
    # срабатывала без какого-либо вменяемого сообщения администратору.
    row = await get_user_record(user_id)
    if not row:
        await callback.answer("❌ Игрок не найден в базе (возможно, был удалён). Обновите список.", show_alert=True)
        return
    nickname = row.get("nickname", "Игрок")

    await assign_tent_db(tent_id, user_id, nickname)
    await sync_tent_row(tent_id)

    tent = await get_tent(tent_id)
    end_date_str = tent[3] if tent else None
    try:
        formatted_end = datetime.strptime(end_date_str, "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        formatted_end = end_date_str or "не установлена"

    # Уведомляем самого игрока о том, что ему выдали палатку — раньше он узнавал об этом
    # только зайдя в бота случайно, никакого сообщения не приходило.
    delivered = await safe_send(
        user_id,
        f"🎉 <b>Вам выдана палатка!</b>\n\n"
        f"⛺ Палатка №{tent_id}\n"
        f"📅 Оплачена до: {formatted_end} 23:59:59\n\n"
        f"Откройте /start → «⛺ Моя палатка», чтобы посмотреть детали или продлить аренду."
    )
    if not delivered:
        logging.error(f"❌ Не удалось уведомить игрока {user_id} о выдаче палатки {tent_id} (возможно, заблокировал бота)")

    await send_log(f"⛺ <b>Выдана палатка:</b>\nПалатка №{tent_id} выдана игроку <code>{nickname}</code> (ID: <code>{user_id}</code>)")
    await callback.answer(f"Палатка №{tent_id} выдана {nickname}!" + ("" if delivered else " (⚠️ уведомить не удалось)"))
    await show_admin_panel(callback.message.chat.id, callback.from_user.id)


@dp.callback_query(F.data.startswith("clear_"))
async def process_clear(callback: types.CallbackQuery):
    if not can_manage_tents(callback.from_user.id):
        return

    _, tent_id, user_id = callback.data.split("_")
    tent_id = int(tent_id)
    tent = await get_tent(tent_id)
    await clear_tent_db(tent_id)
    await sync_tent_row(tent_id)

    if tent and tent[1]:
        try:
            await bot.send_message(
                chat_id=int(user_id),
                text=f"📦 Администратор завершил вашу аренду Палатки №{tent_id}.\n\n⚠️ Пожалуйста, уберите весь ваш товар с палатки в течение 2 дней!"
            )
        except Exception as e:
            logging.error(f"❌ Не удалось уведомить игрока {user_id} об освобождении палатки: {e}")

    await send_log(f"🧹 <b>Освобождена палатка:</b>\nАдминистратор освободил палатку №{tent_id} (была у <code>{tent[2]}</code>)")
    await callback.answer(f"Палатка №{tent_id} освобождена!")
    await show_admin_panel(callback.message.chat.id, callback.from_user.id)


@dp.callback_query(F.data == "adm_users_list")
async def adm_users_list(callback: types.CallbackQuery):
    if not is_super_admin(callback.from_user.id):
        return

    users = await get_all_users()

    if not users:
        await callback.message.edit_text(
            "👥 Зарегистрированных игроков пока нет.",
            reply_markup=InlineKeyboardBuilder().button(text="🔙 Назад", callback_data="back_admin").as_markup()
        )
        return

    msg = "👥 <b>СПИСОК ЗАРЕГИСТРИРУЕМЫХ ИГРОКОВ:</b>\n\n"
    kb = InlineKeyboardBuilder()

    for u_id, u_name, nick in users:
        msg += f"• <b>{nick}</b> (@{u_name})\n"
        kb.button(text=f"✉️ {nick}", callback_data=f"dm_user_{u_id}")
        kb.button(text=f"🚫 Забанить {nick}", callback_data=f"ban_user_{u_id}")
        kb.button(text=f"🗑️ Удалить {nick}", callback_data=f"del_user_{u_id}")

    kb.button(text="📊 Отчёт статистики", callback_data="adm_stats_menu")
    kb.button(text="🔙 Назад в меню", callback_data="back_admin")
    kb.adjust(2)

    await callback.message.edit_text(msg, reply_markup=kb.as_markup(), parse_mode="HTML")


# === УДАЛЕНИЕ ИГРОКА С УКАЗАНИЕМ ПРИЧИНЫ ===
@dp.callback_query(F.data.startswith("del_user_"))
async def process_del_user_start(callback: types.CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        return

    user_id = int(callback.data.split("_")[2])
    row = await get_user_record(user_id)
    if not row:
        await callback.answer("❌ Игрок не найден в базе (возможно, уже удалён). Обновите список.", show_alert=True)
        return
    nick = row.get("nickname", "Игрок")

    await state.update_data(del_user_id=user_id, del_user_nick=nick)
    await state.set_state(DeleteUserState.waiting_for_reason)

    kb = InlineKeyboardBuilder()
    kb.button(text="🗑️ Удалить без указания причины", callback_data="del_user_no_reason")
    kb.button(text="❌ Отменить", callback_data="cancel_state")
    kb.adjust(1)

    await callback.message.answer(
        f"✏️ Укажите причину удаления игрока <b>{nick}</b> следующим сообщением — "
        f"она будет автоматически отправлена ему в личные сообщения.\n\n"
        f"Либо удалите без объяснения причины:",
        reply_markup=kb.as_markup(),
        parse_mode="HTML"
    )
    await callback.answer()


async def finalize_delete_user(chat_id: int, user_id: int, nick: str, reason: str | None):
    await delete_user_db(user_id)
    reason_txt = reason.strip() if reason and reason.strip() else "не указана"

    try:
        await bot.send_message(
            chat_id=user_id,
            text=(
                "🚫 <b>Вы были удалены из базы игроков Торговой Зоны.</b>\n\n"
                f"Причина: {reason_txt}\n\n"
                "Если считаете это ошибкой — свяжитесь с администрацией."
            ),
            parse_mode="HTML"
        )
        delivered = True
    except Exception as e:
        logging.error(f"❌ Не удалось уведомить удалённого игрока {user_id}: {e}")
        delivered = False

    await send_log(f"🗑️ <b>Удаление аккаунта:</b> {nick} (ID: <code>{user_id}</code>)\nПричина: {reason_txt}")
    note = "" if delivered else "\n⚠️ Не удалось доставить уведомление игроку (возможно, он заблокировал бота)."
    await bot.send_message(chat_id=chat_id, text=f"❌ Игрок {nick} удалён из базы.\nПричина: {reason_txt}{note}")


@dp.callback_query(F.data == "del_user_no_reason")
async def process_del_user_no_reason(callback: types.CallbackQuery, state: FSMContext):
    if not is_super_admin(callback.from_user.id):
        return
    data = await state.get_data()
    await state.clear()
    user_id = data.get("del_user_id")
    nick = data.get("del_user_nick", "Игрок")
    if not user_id:
        await callback.answer("⚠️ Сессия удаления устарела, начните заново.", show_alert=True)
        return
    await callback.answer()
    await finalize_delete_user(callback.message.chat.id, user_id, nick, None)


@dp.message(DeleteUserState.waiting_for_reason)
async def process_del_user_reason(message: types.Message, state: FSMContext):
    data = await state.get_data()
    user_id = data.get("del_user_id")
    nick = data.get("del_user_nick", "Игрок")
    await state.clear()
    if not user_id:
        await message.answer("⚠️ Сессия удаления устарела, начните заново через список игроков.")
        return
    await finalize_delete_user(message.chat.id, user_id, nick, message.text)


@dp.callback_query(F.data == "back_admin")
async def back_admin(callback: types.CallbackQuery, state: FSMContext):
    if not can_manage_tents(callback.from_user.id):
        return
    await state.clear()
    await show_admin_panel(callback.message.chat.id, callback.from_user.id)


async def notify_partner_requests():
    """Delivers each new partner request to the owner and administrators once."""
    try:
        requests = await firestore_sync.get_pending_partner_requests()
        for request in requests:
            text = (
                f"🤝 <b>Новая заявка на совладельца палатки №{request.get('tentNum')}</b>\n\n"
                f"Telegram: @{request.get('requesterUsername', '—')}\n"
                f"Minecraft: {request.get('requesterMinecraftNick', '—')}\n"
                f"Заявку можно одобрить или отклонить в Mini App."
            )
            if not request.get("ownerNotifiedAt"):
                delivered = await safe_send(int(request.get("ownerTgId", 0)), text)
                if delivered:
                    await firestore_sync.mark_partner_request_notified(request["id"], "owner")
            if not request.get("adminNotifiedAt"):
                await notify_admins(text)
                await firestore_sync.mark_partner_request_notified(request["id"], "admin")
    except Exception:
        logging.exception("❌ Ошибка уведомления о заявках на совладельца")


@dp.message(Command("claim"))
async def start_tent_claim(message: types.Message, state: FSMContext):
    await state.clear()
    await state.set_state(TentClaimState.waiting_for_tent)
    await message.answer("Укажите номер занятой палатки, к которой хотите получить доступ (1–20):")


@dp.message(TentClaimState.waiting_for_tent)
async def claim_tent_number(message: types.Message, state: FSMContext):
    try:
        tent_num = int(message.text.strip())
    except (TypeError, ValueError):
        await message.answer("Введите номер палатки от 1 до 20.")
        return
    if not 1 <= tent_num <= 20:
        await message.answer("Введите номер палатки от 1 до 20.")
        return
    tent = await firestore_sync.get_tent_data(tent_num)
    if not tent or not tent.get("occupied"):
        await message.answer("Эта палатка свободна или не найдена. Проверьте номер.")
        return
    if tent.get("tgId") and int(tent["tgId"]) == message.from_user.id:
        await state.clear()
        await message.answer("Эта палатка уже привязана к вашему Telegram.")
        return
    await state.update_data(tent_num=tent_num)
    await state.set_state(TentClaimState.waiting_for_nickname)
    await message.answer(f"Введите ваш точный Minecraft-ник для палатки №{tent_num}:")


@dp.message(TentClaimState.waiting_for_nickname)
async def submit_tent_claim(message: types.Message, state: FSMContext):
    nickname = (message.text or "").strip()
    data = await state.get_data()
    tent_num = data.get("tent_num")
    if not nickname or len(nickname) > 32 or not re.fullmatch(r"[A-Za-z0-9_А-Яа-я-]+", nickname):
        await message.answer("Введите Minecraft-ник: до 32 символов, без пробелов.")
        return
    existing = await firestore_sync.get_user_tent(message.from_user.id)
    if existing:
        await state.clear()
        await message.answer(f"У вас уже есть привязанная палатка №{existing[0]}.")
        return
    request_id = await firestore_sync.create_tent_claim({
        "tentNum": tent_num,
        "requesterTgId": message.from_user.id,
        "requesterUsername": message.from_user.username or "",
        "minecraftNick": nickname,
    })
    await state.clear()
    await message.answer("Заявка отправлена администрации. Доступ появится после проверки.")
    await notify_admins(
        f"🔗 <b>Заявка на привязку палатки №{tent_num}</b>\n"
        f"Telegram: @{message.from_user.username or 'без username'} ({message.from_user.id})\n"
        f"Minecraft: {nickname}",
    )


@dp.callback_query(F.data.startswith("claim_"))
async def resolve_tent_claim_callback(callback: types.CallbackQuery):
    if not can_manage_tents(callback.from_user.id):
        await callback.answer("❌ Нет прав.", show_alert=True)
        return
    parts = callback.data.split("_", 2)
    request_id = parts[2]
    approve = parts[1] == "ok"
    try:
        request = await firestore_sync.resolve_tent_claim(request_id, approve, callback.from_user.id)
        if approve:
            text = "✅ Заявка одобрена, палатка привязана к Telegram."
            await safe_send(int(request.get("requesterTgId")), f"✅ Палатка №{request.get('tentNum')} привязана к вашему Telegram.")
        else:
            text = "❌ Заявка отклонена."
            await safe_send(int(request.get("requesterTgId")), f"❌ Заявка на привязку палатки №{request.get('tentNum')} отклонена.")
        await callback.message.edit_text((callback.message.text or "") + f"\n\n{text}")
        await callback.answer(text)
    except ValueError as error:
        await callback.answer(str(error), show_alert=True)


async def notify_tent_claims():
    try:
        for request in await firestore_sync.get_pending_tent_claims():
            if request.get("adminNotifiedAt"):
                continue
            keyboard = InlineKeyboardBuilder()
            keyboard.button(text="✅ Привязать", callback_data=f"claim_ok_{request['id']}")
            keyboard.button(text="❌ Отклонить", callback_data=f"claim_no_{request['id']}")
            keyboard.adjust(2)
            text = (f"🔗 <b>ПРОВЕРКА ВЛАДЕЛЬЦА ПАЛАТКИ №{request.get('tentNum')}</b>\n\n"
                    f"Telegram: @{request.get('requesterUsername') or 'без username'} ({request.get('requesterTgId')})\n"
                    f"Minecraft: {request.get('minecraftNick')}")
            delivered = False
            for admin_id in get_admin_recipients():
                try:
                    await bot.send_message(admin_id, text, reply_markup=keyboard.as_markup(), parse_mode="HTML")
                    delivered = True
                except Exception:
                    logging.exception("Ошибка отправки заявки на привязку админу")
            if delivered:
                await firestore_sync.upsert_bot_document("tent_claim_requests", request["id"], {"adminNotifiedAt": int(time.time() * 1000)})
    except Exception:
        logging.exception("❌ Ошибка уведомления о заявках на привязку")


async def get_current_tariffs() -> dict:
    """Тарифы теперь редактируются администратором прямо в Mini App (коллекция
    Firestore 'tariffs'). Если админ-панель ещё не наполнялась (пустая коллекция),
    используем резервные тарифы из config.py, чтобы бот не сломался."""
    tariffs = await firestore_sync.get_tariffs()
    return tariffs or TARIFFS


async def notify_web_rental_requests():
    """Sends Mini App rental requests to admins; the user never needs the bot chat."""
    try:
        pending = await firestore_sync.get_pending_rental_requests()
        if not pending:
            return
        current_tariffs = await get_current_tariffs()
        for request in pending:
            if request.get("adminNotifiedAt"):
                continue
            tariff = current_tariffs.get(request.get("tariffCode"), {})
            caption = (
                f"📥 <b>ЗАЯВКА ИЗ MINI APP</b>\n\n"
                f"⛺ Палатка №{request.get('tentNum')}\n"
                f"👤 Telegram: @{request.get('username', 'нет')}\n"
                f"🎮 Minecraft: {request.get('minecraftNick', '—')}\n"
                f"📌 Тип: {'Продление' if request.get('requestType') == 'renew' else 'Новая аренда'}\n"
                f"💳 Тариф: {tariff.get('label', request.get('tariffCode', '—'))}"
            )
            kb = InlineKeyboardBuilder()
            kb.button(text="✅ Подтвердить", callback_data=f"webappr_{request['id']}")
            kb.button(text="❌ Отклонить", callback_data=f"webrej_{request['id']}")
            kb.adjust(2)
            await send_request_card_to_admins(request["photoUrl"], caption, kb.as_markup())
            await firestore_sync.mark_rental_request_notified(request["id"])
    except Exception:
        logging.exception("❌ Ошибка отправки заявки аренды из Mini App")


async def notify_chat_messages():
    """Двусторонний чат с администрацией живёт в Mini App (коллекция chat_messages).
    Бот лишь оповещает админов о новых сообщениях от игроков и даёт кнопку «Ответить»,
    которая ведёт в тот же диалог, что и «✉️ Написать игроку» — dm_user_send записывает
    ответ обратно в чат, и игрок видит его прямо в приложении."""
    try:
        for msg in await firestore_sync.get_pending_chat_messages():
            user_id = msg.get("userId")
            username = msg.get("username") or "нет"
            text = (
                f"💬 <b>Новое сообщение в чате Mini App</b>\n\n"
                f"Пользователь: @{username} (ID: <code>{user_id}</code>)\n"
                f"Текст:\n{msg.get('text', '—')}"
            )
            kb = InlineKeyboardBuilder()
            kb.button(text="✉️ Ответить в чате", callback_data=f"dm_user_{user_id}")
            kb.adjust(1)
            for admin_id in get_admin_recipients():
                try:
                    await bot.send_message(chat_id=admin_id, text=text, parse_mode="HTML", reply_markup=kb.as_markup())
                except Exception:
                    pass
            await firestore_sync.mark_chat_message_notified(msg["id"])
    except Exception:
        logging.exception("❌ Ошибка уведомления о сообщении в чате Mini App")


async def notify_license_requests():
    """Заявки на выдачу игровых лицензий, поданные из Mini App."""
    try:
        for request in await firestore_sync.get_pending_license_requests():
            delivery = "\n🚗 Заказана личная доставка (+5 🛢)" if request.get("deliveryRequested") else ""
            text = (
                f"🎫 <b>Новая заявка на лицензию</b>\n\n"
                f"Пользователь: @{request.get('username') or 'нет'} (ID: {request.get('userId')})\n"
                f"Minecraft: {request.get('minecraftNick', '—')}\n"
                f"Сумма: {request.get('totalPrice', 30)} 🛢{delivery}\n"
                f"Комментарий: {request.get('description') or '—'}\n\n"
                f"⚠️ Одобрить/отклонить удобнее прямо в Mini App → Админ-панель → Лицензии — так игрок сразу увидит ответ в чате приложения."
            )
            kb = InlineKeyboardBuilder()
            kb.button(text="✅ Одобрить", callback_data=f"license_appr_{request['id']}")
            kb.button(text="❌ Отклонить", callback_data=f"license_rej_{request['id']}")
            kb.adjust(2)
            for admin_id in get_admin_recipients():
                try:
                    await bot.send_message(chat_id=admin_id, text=text, parse_mode="HTML", reply_markup=kb.as_markup())
                except Exception:
                    pass
            await firestore_sync.mark_license_request_notified(request["id"])
    except Exception:
        logging.exception("❌ Ошибка отправки заявки на лицензию")


@dp.callback_query(F.data.startswith("license_appr_"))
async def approve_license(callback: types.CallbackQuery):
    if not can_manage_tents(callback.from_user.id):
        await callback.answer("❌ Нет прав.", show_alert=True)
        return
    request_id = callback.data.removeprefix("license_appr_")
    request = await firestore_sync.resolve_license_request(request_id, "approved")
    if request and request.get("userId"):
        await safe_send(int(request["userId"]), "✅ Ваша заявка на игровую лицензию одобрена администрацией.")
    try:
        await callback.message.edit_text((callback.message.text or "") + "\n\n✅ ОДОБРЕНО")
    except Exception:
        pass
    await callback.answer("Заявка одобрена")


@dp.callback_query(F.data.startswith("license_rej_"))
async def reject_license(callback: types.CallbackQuery):
    if not can_manage_tents(callback.from_user.id):
        await callback.answer("❌ Нет прав.", show_alert=True)
        return
    request_id = callback.data.removeprefix("license_rej_")
    request = await firestore_sync.resolve_license_request(request_id, "rejected")
    if request and request.get("userId"):
        await safe_send(int(request["userId"]), "❌ Ваша заявка на игровую лицензию отклонена администрацией.")
    try:
        await callback.message.edit_text((callback.message.text or "") + "\n\n❌ ОТКЛОНЕНО")
    except Exception:
        pass
    await callback.answer("Заявка отклонена")


@dp.callback_query(F.data.startswith("webappr_"))
async def approve_web_rental(callback: types.CallbackQuery):
    if not can_manage_tents(callback.from_user.id):
        await callback.answer("❌ У вас нет прав для подтверждения аренды.", show_alert=True)
        return
    request_id = callback.data.removeprefix("webappr_")
    request = await firestore_sync.get_rental_request(request_id)
    if not request or request.get("status") != "pending":
        await callback.answer("⚠️ Заявка уже обработана.", show_alert=True)
        return
    tent_id = int(request["tentNum"])
    tent = await get_tent(tent_id)
    tent_data = await firestore_sync.get_tent_data(tent_id)
    is_renewal = request.get("requestType") == "renew"
    allowed_renewal_user = int(tent_data.get("tgId") or 0) == int(request["userId"]) or int(tent_data.get("partnerTgId") or 0) == int(request["userId"])
    if (is_renewal and (not tent or not allowed_renewal_user)) or (not is_renewal and tent and tent[1] is not None):
        await firestore_sync.resolve_rental_request(request_id, "rejected")
        await callback.answer("⚠️ Палатка уже занята или заявка устарела.", show_alert=True)
        return
    await firestore_sync.resolve_rental_request(request_id, "approved")
    current_tariffs = await get_current_tariffs()
    tariff = current_tariffs.get(request["tariffCode"]) or TARIFFS.get(request["tariffCode"])
    if is_renewal:
        new_end = await extend_tent_db(tent_id, tariff["days"], tariff["price"], request["photoUrl"], request["minecraftNick"])
        end_date = new_end.strftime("%d.%m.%Y")
    else:
        await assign_tent_db(tent_id, int(request["userId"]), request["minecraftNick"], tariff["days"], tariff["price"], request["photoUrl"])
        updated = await get_tent(tent_id)
        end_date = datetime.strptime(updated[3], "%Y-%m-%d").strftime("%d.%m.%Y")
    await sync_tent_row(tent_id)
    receipt_id = f"R-{datetime.now(MSK_TZ).strftime('%d%m%Y-%H%M%S')}"
    receipt = generate_receipt_text(receipt_id, request["minecraftNick"], tent_id, tariff["days"], tariff["price"], end_date)
    await bot.send_message(chat_id=int(request["userId"]), text=receipt, parse_mode="HTML")
    try:
        await callback.message.edit_caption(caption=(callback.message.caption or "") + "\n\n✅ ОДОБРЕНО")
    except Exception:
        pass
    await callback.answer("Аренда подтверждена")


@dp.callback_query(F.data.startswith("webrej_"))
async def reject_web_rental(callback: types.CallbackQuery):
    if not can_manage_tents(callback.from_user.id):
        await callback.answer("❌ У вас нет прав для отклонения заявки.", show_alert=True)
        return
    request_id = callback.data.removeprefix("webrej_")
    request = await firestore_sync.resolve_rental_request(request_id, "rejected")
    if request:
        try:
            await bot.send_message(chat_id=int(request["userId"]), text="❌ Заявка на аренду отклонена администрацией.")
        except Exception:
            pass
    try:
        await callback.message.edit_caption(caption=(callback.message.caption or "") + "\n\n❌ ОТКЛОНЕНО")
    except Exception:
        pass
    await callback.answer("Заявка отклонена")


# === 🚀 ЗАПУСК БОТА И ПЛАНИРОВЩИКА ===
async def sync_all_tents_to_sheets():
    """Периодически «выравнивает» Google-таблицу по текущему состоянию Firestore.

    Раньше строка в Sheets обновлялась только когда бот САМ изменял палатку
    (выдал/снял/продлил). Но палатки теперь редактируются ещё из двух мест —
    Mini App и десктопная программа-менеджер (tz_manager.html) — оба пишут
    напрямую в Firestore, минуя бота. Эта задача читает все 20 палаток из
    единой базы (Firestore) и обновляет реестр, так что таблица всегда
    отражает реальное состояние независимо от того, кто внёс изменение.

    ВАЖНО: коллекция bot_payments читается ОДИН раз за весь проход (а не 20 раз —
    по одному на каждую палатку), иначе это моментально выжигает бесплатную
    дневную квоту чтений Firestore."""
    if not sheets_sync.is_configured():
        return
    try:
        all_payments = await firestore_sync.list_bot_documents("rental_history")
        payments_by_tent = {}
        for item in all_payments:
            payments_by_tent.setdefault(int(item.get("tentNum", 0)), []).append({
                "price": item.get("amount", 0),
                "operationAt": item.get("operationAt", 0),
            })
        for tid in range(1, 21):
            await sync_tent_row(tid, payments_by_tent=payments_by_tent)
            await asyncio.sleep(1.1)  # не упереться в лимит Google Sheets API
    except Exception as e:
        logging.error(f"❌ Ошибка периодической синхронизации палаток с Google Sheets: {e}")


async def main():
    if MINIAPP_URL:
        try:
            await bot.set_chat_menu_button(
                menu_button=types.MenuButtonWebApp(text="Приложение", web_app=types.WebAppInfo(url=MINIAPP_URL))
            )
        except Exception as e:
            logging.error(f"❌ Не удалось установить кнопку меню Mini App: {e}")

    scheduler = AsyncIOScheduler(timezone=MSK_TZ)
    scheduler.add_job(check_tent_expirations, 'cron', hour=12, minute=0)
    scheduler.add_job(products_sync.sync_products, 'interval', minutes=products_sync.sync_interval_minutes(), next_run_time=datetime.now())
    scheduler.add_job(notify_partner_requests, 'interval', minutes=2, next_run_time=datetime.now())
    scheduler.add_job(notify_web_rental_requests, 'interval', minutes=2, next_run_time=datetime.now())
    scheduler.add_job(notify_tent_claims, 'interval', minutes=2, next_run_time=datetime.now())
    scheduler.add_job(notify_chat_messages, 'interval', seconds=45, next_run_time=datetime.now())
    scheduler.add_job(notify_license_requests, 'interval', minutes=2, next_run_time=datetime.now())
    scheduler.add_job(sync_all_tents_to_sheets, 'interval', minutes=20)
    scheduler.add_job(send_db_backup, 'interval', days=3)
    scheduler.start()

    # Одноразовая безопасная миграция: досоздаёт в Firestore только те палатки, которых
    # там ещё вообще нет. Если документ для номера палатки уже существует — он считается
    # авторитетным (например, уже введён через веб-приложение) и НЕ перезаписывается.
    if firestore_sync.is_configured():
        try:
            report = await firestore_sync.migrate_sqlite_if_missing()
            if report.get("no_firestore"):
                logging.error("❌ Firestore недоступен при старте — проверьте ключ FIREBASE_SERVICE_ACCOUNT_FILE и доступ в интернет.")
            elif report.get("migrated"):
                await send_log(
                    f"🔄 <b>Миграция в Firestore при старте:</b>\n"
                    f"Досозданы палатки: {report['migrated']}\n"
                    f"Уже существовали (не тронуты): {report['skipped_existing'] or '—'}"
                )
        except Exception as e:
            logging.error(f"❌ Ошибка миграции SQLite → Firestore при старте: {e}")
    else:
        logging.warning("⚠️ FIREBASE_SERVICE_ACCOUNT_FILE не настроен — бот не сможет читать/писать палатки!")

    if sheets_sync.is_configured():
        try:
            await sync_all_tents_to_sheets()
        except Exception as e:
            logging.error(f"❌ Не удалось выполнить стартовую синхронизацию с Google Sheets: {e}")

    await send_log("🟢 <b>Бот успешно запущен и готов к работе!</b>")
    print("🤖 Бот запущен! Планировщик (бэкапы и напоминания) активен.")
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
